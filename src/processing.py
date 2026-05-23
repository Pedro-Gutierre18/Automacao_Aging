from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

import pandas as pd

def ler_arquivos():
    export_bi = pd.read_excel(
        r"data\modelo\export.xlsx",
        engine="openpyxl",
        sheet_name="Export",
        skipfooter=3,
        )

    ultima_linha = len(export_bi) + 1

    # Lê o Modelo Aging e define a planilha ativa - Destino
    aging = load_workbook(r"data\modelo\Modelo Aging.xlsx")
    ws_aging = aging["Export"]

    # Pega o nome da Tabela
    tabela = ws_aging.tables['Tabela']

    return {
        "export_bi": export_bi,
        "aging": aging,
        "ultima_linha": ultima_linha,
        "ws_aging": ws_aging,
        "tabela": tabela
        }

def tratamento(dados, data_ontem, data_bi):
    # Dicionario com as variaveis
    export_bi   = dados["export_bi"]
    aging       = dados["aging"]
    ultima_linha = dados["ultima_linha"]
    ws_aging    = dados["ws_aging"]
    tabela      = dados["tabela"]

    # Importar os dados para o arquivo modelo
    for i, linha in enumerate(export_bi.itertuples(index=False), start=2):
        for j, valor in enumerate(linha, start=1):
            ws_aging.cell(row=i, column=j, value=valor)

    # Reaplica as fórmulas usadas no modelo
    formulas = {
        "R": lambda r: f"=VLOOKUP(K{r},i!$A$2:$B$47,2,FALSE)",
        "S": lambda r: f"=VLOOKUP(K{r},'Etapas da conta'!$F$4:$G$28,2,FALSE)",
        "T": lambda r: f"=IF(A{r}=\"SADT\",\"INTERNADO\",A{r})",
        "V": lambda r: f"=IF(_xlfn.XLOOKUP(Z{r},'Baixa contabil'!$N$2:$N$29,'Baixa contabil'!$N$2:$N$29,\"\",0)<>\"\",\"SIM\",\"NÃO\")",
        "W": lambda r: (
            f"=IF(ISNUMBER(SEARCH(\"FAT - AGUARDA ENVIO XML [E EMISSÃO RPS]\",K{r})),\"\","
            f"IF(ISNUMBER(SEARCH(\"FAT - AGUARDA ENVIO XML E EMISSÃO RPS\",K{r})),\"\","
            f"IF(ISNUMBER(SEARCH(\"TESOURARIA - BAIXA ADMINISTRATIVA\",K{r})),\"\",K{r})))"),
        "X": lambda r: f"=IF(O{r}=\"Autorizado\",\"Autorizado\",\"Pendente\")",
        "Y": lambda r: f"=IFERROR(IF(AND(_xlfn.XLOOKUP(E{r},i!$D$2:$D$52,i!$G$2:$G$52)<>0,J{r}<>\"\"),J{r}+_xlfn.XLOOKUP(E{r},i!$D$2:$D$52,i!$G$2:$G$52),\"\"),\"\")",
        "Z": lambda r: f"=IF(_xlfn.XLOOKUP(E{r},i!$D$2:$D$52,i!$H$2:$H$52,\"\")<>0,_xlfn.XLOOKUP(E{r},i!$D$2:$D$52,i!$H$2:$H$52,\"\"),\"\")",
        "AA": lambda r: f"=_xlfn.CONCAT(C{r},\"-\",F{r})"
    }

    for col_letra, formula_fn in formulas.items():
        for row_num in range(2, ultima_linha + 1):
            ws_aging[f"{col_letra}{row_num}"] = formula_fn(row_num)

    # Formata a coluna 'Valor Total' para o formato de moeda
    for cell in ws_aging['Q'][1:]:
        cell.number_format = "R$ #,##0.00"

    # Formatando colunas de data
    colunas_data = ['G', 'H', 'I', 'J', 'Y']

    for coluna in colunas_data:    
        for cell in ws_aging[coluna][1:]:
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = "DD/MM/YYYY"

    # Alinhando outras colunas
    alinhar_direita = ['B', 'C', 'Q']

    for coluna in alinhar_direita:
        for cell in ws_aging[coluna][1:]:
            cell.alignment = Alignment(horizontal='right')

    alinhar_centro = ['F', 'AA']

    for coluna in alinhar_centro:
        for cell in ws_aging[coluna][1:]:
            cell.alignment = Alignment(horizontal='center')

    # Aplicando cores em aloggingumas colunas
    for cell in ws_aging['AA'][1:]:
        if cell.value is not None:
            cell.fill = PatternFill(fill_type='solid', fgColor='61CBF3')
            cell.font = Font(color='002060', bold=True)

    for cell in ws_aging['R'][1:]:
        if cell.value is not None:
            cell.font = Font(color='FF0000')

    col_auxiliares = ['S', 'T', 'U', 'V', 'W', 'X']

    for coluna in col_auxiliares:
        for cell in ws_aging[coluna][1:ultima_linha]:
                cell.fill = PatternFill(fill_type='solid', fgColor='F2F2F2')

    # Atualiza o intervalo para evitar dados fora da tabela
    ultima_coluna = get_column_letter(ws_aging.max_column)
    tabela.ref = f"A1:{ultima_coluna}{ultima_linha}"

    # Adicionando data na página da Tabela Dinâmica
    pivot = aging["Aging (Etapas)"]
    pivot["C1"] = data_ontem
    pivot["E1"] = data_bi


def salvar_arquivo(data_aging, aging):
    # Salvando arquivo tratado
    aging.save(rf"resultado\{data_aging}_Aging.xlsx")
